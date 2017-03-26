# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import psycopg2

# 테이블을 두개로 나누자
# archive_research table -> 팀 ID Key로 연구 논문 데이터만 저장
# archive_researcher table -> 유저 ID Key로 연구 참여자 데이터 저장
# 연구 참가자 데이터 추출 -> left join

# A     B  (U_K) C   D       E       F   G   H       I       J       K      M   N     O
# 팀 ID	개인 ID	이름	이메일	연락처	소속	학년	참가유형	논문제목	논문초록	논문파일	분야	입력한 팀원수


MAX_ROWS = 529
MAX_COLS = 13

wb = load_workbook('./presentation_output.xlsx')
ws = wb.active

COL_NAMES = {}
research_datas = []
researcher_datas = []
# Parse col names
for i in range(1, MAX_COLS+2):
    print('col_name i: ', i)
    COL_NAMES[i] = ws.cell(row=1, column=i).value

# print COL_NAMES
for i in range(1, len(COL_NAMES)):
    print(COL_NAMES[i])

print('COLNAME[14]: ', COL_NAMES[14])
for i in range(2, MAX_ROWS):
    research_data = {}
    researcher_data = {}

    if '데이터 수동 확인 필요' not in ws.cell(row=i, column=3).value:

        # 입력한 팀원 수 가 None일 경우, researcher_data에만 추가함
        if ws.cell(row=i, column=MAX_COLS).value is not None:
            for j in range(1, len(COL_NAMES)):
                research_data[COL_NAMES[j]] = ws.cell(row=i, column=j).value
            research_datas.append(research_data)
        # 모든 데이터는 기본적으로 researcher_data에 추가
        # 팀 ID
        # 개인 ID
        # 이름
        # 이메일
        # 연락처
        # 소속
        # 학년
        # 참가 유형
        print('adding researcher_data : ', i)
        for j in range(1, 8):
            researcher_data[COL_NAMES[j]] = ws.cell(row=i, column=j).value
        researcher_datas.append(researcher_data)
    else:
        print('Invalid data ')

for i in range(1, len(research_datas)):
    print(research_datas[i])

for i in range(1, len(researcher_datas)):
    print(researcher_datas[i])

print(researcher_datas[0][COL_NAMES[3]])
# 데이터베이스에 입력
conn = psycopg2.connect("host='localhost' dbname='yscholar' user='yscholar' password='secret'")

# 커서를 연다
cur = conn.cursor()

# 연구 데이터 입력
sql_research = "INSERT INTO archive_research (team_id, researcher_id, researcher_name, email, phone_number, org, grade, type, research_name, abstract_kor, file_url, category, team_num) VALUES (%s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
for i in range(1, len(research_datas)):
    data = research_datas[i]
    cur.execute(sql_research, (
    data[COL_NAMES[1]], data[COL_NAMES[2]], data[COL_NAMES[3]], data[COL_NAMES[4]], data[COL_NAMES[5]],
    data[COL_NAMES[6]], data[COL_NAMES[7]], data[COL_NAMES[8]], data[COL_NAMES[9]], data[COL_NAMES[10]],
    data[COL_NAMES[11]], data[COL_NAMES[12]], data[COL_NAMES[13]]))

# 연구자 데이터 입력
sql_researcher = "INSERT INTO archive_researcher (team_id, researcher_id, researcher_name, email, phone_number, org, grade) VALUES (%s, %s,%s, %s, %s, %s, %s)"
for i in range(1, len(researcher_datas)):
    data = researcher_datas[i]
    cur.execute(sql_researcher, (data[COL_NAMES[1]], data[COL_NAMES[2]], data[COL_NAMES[3]], data[COL_NAMES[4]], data[COL_NAMES[5]],
    data[COL_NAMES[6]], data[COL_NAMES[7]]))

#cur.execute()
# 데이터를 수정했을 경우 반드시 commit
conn.commit()

# 연결을 종료한다
cur.close()
conn.close()
